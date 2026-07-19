#!/usr/bin/env python3
import base64
import cgi
import datetime as dt
import hashlib
import hmac
import html
import io
import json
import os
import re
import secrets
import shutil
import sqlite3
import sys
import traceback
import uuid
import zipfile
from decimal import Decimal, ROUND_HALF_UP
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


APP_NAME = "BL Invoice Generation"
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("BLINVOICE_DATA_DIR", BASE_DIR / "data")).resolve()
DB_PATH = Path(os.environ.get("BLINVOICE_DB", DATA_DIR / "blinvoice.sqlite3")).resolve()
UPLOAD_DIR = DATA_DIR / "uploads"
INVOICE_DIR = DATA_DIR / "invoices"
SESSION_DAYS = 1
EXPECTED_HEADERS = [
    "InvoiceDate",
    "InvoiceMonth",
    "BankingLabsLocation",
    "Customer",
    "ResourceName",
    "Task",
    "Currency",
    "Rate",
    "Hours",
    "InvoiceType",
    "LTO",
    "PONumber",
]


def ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    INVOICE_DIR.mkdir(parents=True, exist_ok=True)


def get_secret():
    ensure_dirs()
    secret_path = DATA_DIR / "secret.key"
    if not secret_path.exists():
        secret_path.write_text(secrets.token_urlsafe(48), encoding="utf-8")
        os.chmod(secret_path, 0o600)
    return secret_path.read_text(encoding="utf-8").strip().encode("utf-8")


def db():
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def utcnow():
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def parse_date(value):
    if value is None or value == "":
        return dt.date.today()
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Invalid date: {text}. Use dd/mm/yyyy or yyyy-mm-dd.")


def parse_month(value):
    if value is None or value == "":
        today = dt.date.today()
        return today.year, today.month
    if isinstance(value, (int, float)):
        text = str(int(value))
    else:
        text = str(value).strip()
    if len(text) == 6 and text.isdigit():
        year, month = int(text[:4]), int(text[4:])
        if 1 <= month <= 12:
            return year, month
    for fmt in ("%Y-%m", "%m/%Y", "%b %Y", "%B %Y"):
        try:
            d = dt.datetime.strptime(text, fmt)
            return d.year, d.month
        except ValueError:
            pass
    raise ValueError(f"Invalid invoice month: {text}. Use yyyymm, yyyy-mm, or mmm yyyy.")


def month_range_text(year, month):
    first = dt.date(year, month, 1)
    if month == 12:
        last = dt.date(year + 1, 1, 1) - dt.timedelta(days=1)
    else:
        last = dt.date(year, month + 1, 1) - dt.timedelta(days=1)
    return f"{first:%d/%m/%Y}-{last:%d/%m/%Y}"


def money(value):
    return Decimal(str(value or "0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def fmt_money(value):
    return f"{money(value):,.2f}"


def fmt_date(value):
    if isinstance(value, str):
        value = dt.date.fromisoformat(value)
    return value.strftime("%d/%m/%Y")


def sanitize_filename_part(value):
    return re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").strip()).strip("_") or "Unknown"


def invoice_pdf_filename(resource_names, invoice_month, invoice_number):
    resource = sanitize_filename_part(resource_names)
    month = sanitize_filename_part(invoice_month)
    return f"Invoice_{resource}_{month}_{invoice_number}.pdf"


def normalize_key(value):
    return str(value or "").strip().casefold()


def pbkdf2_hash(password, salt=None):
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 260000)
    return "pbkdf2_sha256$260000$%s$%s" % (
        base64.b64encode(salt).decode("ascii"),
        base64.b64encode(digest).decode("ascii"),
    )


def check_password(password, stored):
    try:
        algo, rounds, salt_b64, digest_b64 = stored.split("$", 3)
        if algo != "pbkdf2_sha256":
            return False
        salt = base64.b64decode(salt_b64)
        expected = base64.b64decode(digest_b64)
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(rounds))
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False


def init_db():
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'admin',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS locations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                address_line1 TEXT NOT NULL,
                address_line2 TEXT NOT NULL DEFAULT '',
                phone TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL DEFAULT '',
                website TEXT NOT NULL DEFAULT '',
                hst_number TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                external_code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                address_line1 TEXT NOT NULL,
                address_line2 TEXT NOT NULL DEFAULT '',
                city TEXT NOT NULL DEFAULT '',
                province_postal TEXT NOT NULL DEFAULT '',
                country TEXT NOT NULL DEFAULT '',
                attention TEXT NOT NULL DEFAULT '',
                default_po TEXT NOT NULL DEFAULT '',
                active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS tax_rates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                percent REAL NOT NULL,
                active INTEGER NOT NULL DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS invoice_sequence (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                next_number INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS upload_batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_uuid TEXT NOT NULL UNIQUE,
                filename TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                uploaded_by INTEGER NOT NULL,
                uploaded_at TEXT NOT NULL,
                status TEXT NOT NULL,
                row_count INTEGER NOT NULL DEFAULT 0,
                invoice_count INTEGER NOT NULL DEFAULT 0,
                error_json TEXT NOT NULL DEFAULT '[]',
                FOREIGN KEY(uploaded_by) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER NOT NULL,
                invoice_number INTEGER NOT NULL UNIQUE,
                invoice_date TEXT NOT NULL,
                due_date TEXT NOT NULL,
                terms TEXT NOT NULL,
                location_id INTEGER NOT NULL,
                customer_id INTEGER NOT NULL,
                currency TEXT NOT NULL,
                po_number TEXT NOT NULL,
                lto TEXT NOT NULL DEFAULT '',
                subtotal REAL NOT NULL,
                tax_percent REAL NOT NULL,
                tax_amount REAL NOT NULL,
                total REAL NOT NULL,
                pdf_path TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(batch_id) REFERENCES upload_batches(id),
                FOREIGN KEY(location_id) REFERENCES locations(id),
                FOREIGN KEY(customer_id) REFERENCES customers(id)
            );

            CREATE TABLE IF NOT EXISTS invoice_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER NOT NULL,
                source_row INTEGER NOT NULL,
                invoice_month TEXT NOT NULL,
                activity TEXT NOT NULL,
                resource_name TEXT NOT NULL,
                description TEXT NOT NULL,
                invoice_type TEXT NOT NULL,
                hours REAL NOT NULL,
                rate REAL NOT NULL,
                amount REAL NOT NULL,
                raw_json TEXT NOT NULL,
                FOREIGN KEY(invoice_id) REFERENCES invoices(id)
            );
            """
        )
        if not conn.execute("SELECT 1 FROM invoice_sequence WHERE id = 1").fetchone():
            conn.execute("INSERT INTO invoice_sequence (id, next_number) VALUES (1, 51685)")
        if not conn.execute("SELECT 1 FROM users").fetchone():
            conn.execute(
                "INSERT INTO users (username, password_hash, role, active, created_at) VALUES (?, ?, 'admin', 1, ?)",
                ("admin", pbkdf2_hash("admin123"), utcnow()),
            )
        seed_master_data(conn)


def seed_master_data(conn):
    if not conn.execute("SELECT 1 FROM locations").fetchone():
        conn.execute(
            """
            INSERT INTO locations
            (code, name, address_line1, address_line2, phone, email, website, hst_number, active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            """,
            (
                "250Y",
                "Banking Labs - 250 Yonge",
                "250 Yonge St",
                "Suite # 2201 Toronto Ontario M5B 2L7",
                "6479641268",
                "Connect@bankinglabs.com",
                "https://bankinglabs.com/",
                "821806635RT0001",
            ),
        )
    customers = [
        ("2", "Bank of Montreal", "100 King Street West", "Floor - 11", "Toronto", "Ontario M5X 1A1", "Canada", "Maria Cristina Monica DULOS", ""),
        ("4", "Bank of Montreal", "100 King Street West", "Floor - 11", "Toronto", "Ontario M5X 1A1", "Canada", "Maria Cristina Monica DULOS", "31806438"),
    ]
    for row in customers:
        if not conn.execute("SELECT 1 FROM customers WHERE external_code = ?", (row[0],)).fetchone():
            conn.execute(
                """
                INSERT INTO customers
                (external_code, name, address_line1, address_line2, city, province_postal, country, attention, default_po, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """,
                row,
            )
    for task in ("Consulting",):
        if not conn.execute("SELECT 1 FROM tasks WHERE name = ?", (task,)).fetchone():
            conn.execute("INSERT INTO tasks (name, active) VALUES (?, 1)", (task,))
    if not conn.execute("SELECT 1 FROM tax_rates WHERE code = 'HST'").fetchone():
        conn.execute("INSERT INTO tax_rates (code, percent, active) VALUES ('HST', 13.0, 1)")


def sign_session(user_id):
    expires = int((dt.datetime.now(dt.timezone.utc) + dt.timedelta(days=SESSION_DAYS)).timestamp())
    payload = f"{user_id}:{expires}"
    sig = hmac.new(get_secret(), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{payload}:{sig}"


def verify_session(token):
    if not token:
        return None
    try:
        user_id, expires, sig = token.split(":", 2)
        payload = f"{user_id}:{expires}"
        expected = hmac.new(get_secret(), payload.encode("utf-8"), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected):
            return None
        if int(expires) < int(dt.datetime.now(dt.timezone.utc).timestamp()):
            return None
        with db() as conn:
            return conn.execute(
                "SELECT id, username, role FROM users WHERE id = ? AND active = 1", (int(user_id),)
            ).fetchone()
    except Exception:
        return None


def esc(value):
    return html.escape("" if value is None else str(value), quote=True)


def money_words(amount):
    amount = money(amount)
    dollars = int(amount)
    cents = int((amount - dollars) * 100)
    words = f"{number_words(dollars).title()} Dollars"
    if cents:
        words += f" and {number_words(cents).title()} Cents"
    return words


ONES = [
    "zero",
    "one",
    "two",
    "three",
    "four",
    "five",
    "six",
    "seven",
    "eight",
    "nine",
    "ten",
    "eleven",
    "twelve",
    "thirteen",
    "fourteen",
    "fifteen",
    "sixteen",
    "seventeen",
    "eighteen",
    "nineteen",
]
TENS = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]


def number_words(n):
    n = int(n)
    if n < 20:
        return ONES[n]
    if n < 100:
        return TENS[n // 10] + ("" if n % 10 == 0 else " " + ONES[n % 10])
    if n < 1000:
        return ONES[n // 100] + " hundred" + ("" if n % 100 == 0 else " " + number_words(n % 100))
    if n < 1_000_000:
        return number_words(n // 1000) + " thousand" + ("" if n % 1000 == 0 else " " + number_words(n % 1000))
    return number_words(n // 1_000_000) + " million" + ("" if n % 1_000_000 == 0 else " " + number_words(n % 1_000_000))


def get_master_maps(conn):
    locations = conn.execute("SELECT * FROM locations WHERE active = 1 ORDER BY code").fetchall()
    customers = conn.execute("SELECT * FROM customers WHERE active = 1 ORDER BY external_code").fetchall()
    tasks = conn.execute("SELECT * FROM tasks WHERE active = 1 ORDER BY name").fetchall()
    tax = conn.execute("SELECT * FROM tax_rates WHERE active = 1 ORDER BY id LIMIT 1").fetchone()
    loc_map = {normalize_key(x["code"]): x for x in locations}
    cust_map = {}
    for c in customers:
        cust_map[normalize_key(c["external_code"])] = c
        cust_map[normalize_key(c["name"])] = c
        cust_map[normalize_key(f"{c['external_code']} - {c['name']}")] = c
    task_map = {normalize_key(x["name"]): x for x in tasks}
    return locations, customers, tasks, tax, loc_map, cust_map, task_map


def read_excel_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ValueError("Excel headers do not match expected template: " + ", ".join(EXPECTED_HEADERS))
    rows = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        rows.append((idx, dict(zip(EXPECTED_HEADERS, values))))
    if not rows:
        raise ValueError("Uploaded Excel does not contain invoice rows.")
    return rows


def next_invoice_number(conn):
    row = conn.execute("SELECT next_number FROM invoice_sequence WHERE id = 1").fetchone()
    current = int(row["next_number"])
    conn.execute("UPDATE invoice_sequence SET next_number = ? WHERE id = 1", (current + 1,))
    return current


def create_invoices_from_upload(upload_path, original_name, user_id):
    batch_uuid = str(uuid.uuid4())
    stored = UPLOAD_DIR / f"{batch_uuid}_{Path(original_name).name}"
    shutil.copyfile(upload_path, stored)
    errors = []
    parsed = []
    with db() as conn:
        locations, customers, tasks, tax, loc_map, cust_map, task_map = get_master_maps(conn)
        tax_percent = Decimal(str(tax["percent"] if tax else 13))
        rows = read_excel_rows(stored)
        for source_row, row in rows:
            try:
                invoice_date = parse_date(row["InvoiceDate"])
                due_date = invoice_date + dt.timedelta(days=30)
                year, month = parse_month(row["InvoiceMonth"])
                location = loc_map.get(normalize_key(row["BankingLabsLocation"]))
                customer = cust_map.get(normalize_key(row["Customer"]))
                task_name = str(row["Task"] or "").strip()
                task = task_map.get(normalize_key(task_name))
                if not location:
                    raise ValueError(f"Unknown BankingLabsLocation '{row['BankingLabsLocation']}'")
                if not customer:
                    raise ValueError(f"Unknown Customer '{row['Customer']}'")
                if not task:
                    raise ValueError(f"Unknown Task '{task_name}'")
                rate = money(row["Rate"])
                hours = Decimal(str(row["Hours"] or "0")).quantize(Decimal("0.01"))
                amount = money(rate * hours)
                currency = str(row["Currency"] or "CAD").strip().upper()
                invoice_type = str(row["InvoiceType"] or "Hours").strip()
                po_number = str(row["PONumber"] or customer["default_po"] or "-").strip()
                lto = str(row["LTO"] or "").strip()
                resource = str(row["ResourceName"] or "").strip()
                if not resource:
                    raise ValueError("ResourceName is required")
                parsed.append(
                    {
                        "source_row": source_row,
                        "row": row,
                        "invoice_date": invoice_date,
                        "due_date": due_date,
                        "invoice_month": f"{year:04d}{month:02d}",
                        "duration": month_range_text(year, month),
                        "location": location,
                        "customer": customer,
                        "task": task_name,
                        "currency": currency,
                        "rate": rate,
                        "hours": hours,
                        "amount": amount,
                        "invoice_type": invoice_type,
                        "po_number": po_number,
                        "lto": lto,
                        "resource": resource,
                    }
                )
            except Exception as exc:
                errors.append({"row": source_row, "error": str(exc)})
        status = "failed" if errors else "completed"
        cur = conn.execute(
            """
            INSERT INTO upload_batches
            (batch_uuid, filename, stored_path, uploaded_by, uploaded_at, status, row_count, invoice_count, error_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, 0, ?)
            """,
            (batch_uuid, original_name, str(stored), user_id, utcnow(), status, len(rows), json.dumps(errors)),
        )
        batch_id = cur.lastrowid
        invoice_ids = []
        if not errors:
            groups = {}
            for item in parsed:
                key = (
                    item["invoice_date"].isoformat(),
                    item["location"]["id"],
                    item["customer"]["id"],
                    item["currency"],
                    item["po_number"],
                    item["lto"],
                )
                groups.setdefault(key, []).append(item)
            for _, items in groups.items():
                invoice_number = next_invoice_number(conn)
                first = items[0]
                subtotal = sum((x["amount"] for x in items), Decimal("0.00"))
                tax_amount = money(subtotal * tax_percent / Decimal("100"))
                total = money(subtotal + tax_amount)
                pdf_path = INVOICE_DIR / f"invoice_{invoice_number}.pdf"
                conn.execute(
                    """
                    INSERT INTO invoices
                    (batch_id, invoice_number, invoice_date, due_date, terms, location_id, customer_id,
                     currency, po_number, lto, subtotal, tax_percent, tax_amount, total, pdf_path, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id,
                        invoice_number,
                        first["invoice_date"].isoformat(),
                        first["due_date"].isoformat(),
                        "Net 30 Days",
                        first["location"]["id"],
                        first["customer"]["id"],
                        first["currency"],
                        first["po_number"],
                        first["lto"],
                        float(subtotal),
                        float(tax_percent),
                        float(tax_amount),
                        float(total),
                        str(pdf_path),
                        utcnow(),
                    ),
                )
                invoice_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                invoice_ids.append(invoice_id)
                for item in items:
                    description = f"{item['resource']} Duration\n{item['duration']}"
                    conn.execute(
                        """
                        INSERT INTO invoice_items
                        (invoice_id, source_row, invoice_month, activity, resource_name, description,
                         invoice_type, hours, rate, amount, raw_json)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            invoice_id,
                            item["source_row"],
                            item["invoice_month"],
                            item["task"],
                            item["resource"],
                            description,
                            item["invoice_type"],
                            float(item["hours"]),
                            float(item["rate"]),
                            float(item["amount"]),
                            json.dumps({k: str(v) if v is not None else "" for k, v in item["row"].items()}),
                        ),
                    )
                render_invoice_pdf(conn, invoice_id)
            conn.execute(
                "UPDATE upload_batches SET invoice_count = ? WHERE id = ?",
                (len(invoice_ids), batch_id),
            )
    return batch_id, errors, invoice_ids


def draw_wrapped(c, text, x, y, width, font="Helvetica", size=10, leading=12):
    c.setFont(font, size)
    lines = []
    for raw_line in str(text or "").splitlines():
        words = raw_line.split()
        current = ""
        for word in words:
            test = word if not current else current + " " + word
            if stringWidth(test, font, size) <= width:
                current = test
            else:
                if current:
                    lines.append(current)
                current = word
        lines.append(current)
    for line in lines:
        c.drawString(x, y, line)
        y -= leading
    return y


def render_invoice_pdf(conn, invoice_id):
    inv = conn.execute(
        """
        SELECT i.*, l.*, c.external_code AS customer_code, c.name AS customer_name,
               c.address_line1 AS customer_address1, c.address_line2 AS customer_address2,
               c.city AS customer_city, c.province_postal AS customer_province_postal,
               c.country AS customer_country, c.attention AS customer_attention,
               l.name AS location_name
        FROM invoices i
        JOIN locations l ON l.id = i.location_id
        JOIN customers c ON c.id = i.customer_id
        WHERE i.id = ?
        """,
        (invoice_id,),
    ).fetchone()
    items = conn.execute("SELECT * FROM invoice_items WHERE invoice_id = ? ORDER BY id", (invoice_id,)).fetchall()
    path = Path(inv["pdf_path"])
    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    margin_left = 49
    blue = colors.HexColor("#4a78d4")
    header_fill = colors.HexColor("#d7e0f2")

    def page_header():
        c.setFillColor(colors.HexColor("#50b36b"))
        c.rect(80, 772, 17, 8, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#f2ce36"))
        c.rect(80, 760, 17, 8, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#eb3b32"))
        c.rect(80, 748, 17, 8, fill=1, stroke=0)
        c.setFillColor(blue)
        c.setFont("Helvetica-Bold", 52)
        c.drawString(96, 741, "B")
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 26)
        c.drawString(141, 756, "Banking Labs")

        c.setFont("Helvetica", 10)
        c.drawString(366, 765, inv["address_line1"])
        c.drawString(366, 754, inv["address_line2"])
        c.drawString(366, 743, f"Phone: {inv['phone']}")
        c.drawString(366, 732, f"Email: {inv['email']}")
        c.drawString(366, 720, inv["website"])

        c.setFillColor(blue)
        c.setFont("Helvetica", 18)
        c.drawString(margin_left, 686, "INVOICE")
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin_left, 665, "INVOICE TO")
        c.setFont("Helvetica", 10)
        y = 638
        for value in (
            inv["customer_name"],
            inv["customer_address1"],
            inv["customer_address2"],
            inv["customer_city"],
            inv["customer_province_postal"],
            inv["customer_country"],
        ):
            if value:
                c.drawString(margin_left, y, value)
                y -= 13
        if inv["customer_attention"]:
            c.drawString(margin_left, y - 20, f"Attn: {inv['customer_attention']}")

        x_label, x_value = 366, 420
        meta_y = 665
        meta = [
            ("Invoice #", str(inv["invoice_number"])),
            ("Date", fmt_date(inv["invoice_date"])),
            ("Due Date", fmt_date(inv["due_date"])),
            ("Terms", inv["terms"]),
            ("PO #", inv["po_number"]),
        ]
        c.setFont("Helvetica", 10)
        for label, value in meta:
            c.drawString(x_label, meta_y, label)
            c.drawString(x_value, meta_y, value)
            meta_y -= 22
        if inv["lto"]:
            c.drawString(x_label, meta_y, "LTO")
            c.drawString(x_value, meta_y, inv["lto"])

        c.setStrokeColor(blue)
        c.setLineWidth(1.4)
        c.line(74, 500, 525, 500)
        c.setStrokeColor(colors.black)

    page_header()
    table_x = 42
    table_y = 480
    row_h = 90
    header_h = 21
    col_widths = [68, 75, 116, 41, 41, 55, 54, 56]
    labels = ["Date", "Activity", "Description", "Hours", "Rate", "Amount", "Tax (%)", "Total"]
    x = table_x
    c.setFillColor(header_fill)
    c.rect(table_x, table_y - header_h, sum(col_widths), header_h, fill=1, stroke=0)
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.7)
    for w in col_widths:
        c.rect(x, table_y - header_h - row_h * max(1, len(items)), w, header_h + row_h * max(1, len(items)), fill=0, stroke=1)
        x += w
    c.line(table_x, table_y - header_h, table_x + sum(col_widths), table_y - header_h)
    x = table_x
    c.setFillColor(blue)
    c.setFont("Helvetica-Bold", 10)
    for label, w in zip(labels, col_widths):
        c.drawString(x + 5, table_y - 13, label)
        x += w
    c.setFillColor(colors.black)
    y = table_y - header_h - 16
    for item in items:
        x = table_x
        c.setFont("Helvetica", 10)
        c.drawString(x + 3, y, fmt_date(inv["invoice_date"]))
        x += col_widths[0]
        c.drawString(x + 3, y, item["activity"])
        x += col_widths[1]
        draw_wrapped(c, item["description"], x + 3, y, col_widths[2] - 6, size=10, leading=12)
        x += col_widths[2]
        c.drawRightString(x + col_widths[3] - 4, y, fmt_money(item["hours"]))
        x += col_widths[3]
        c.drawRightString(x + col_widths[4] - 4, y, fmt_money(item["rate"]))
        x += col_widths[4]
        c.drawRightString(x + col_widths[5] - 4, y, fmt_money(item["amount"]))
        x += col_widths[5]
        c.drawRightString(x + col_widths[6] - 4, y, fmt_money(inv["tax_percent"]))
        x += col_widths[6]
        item_total = money(Decimal(str(item["amount"])) * (Decimal("1") + Decimal(str(inv["tax_percent"])) / Decimal("100")))
        c.drawRightString(x + col_widths[7] - 4, y, fmt_money(item_total))
        y -= row_h
    table_bottom = table_y - header_h - row_h * max(1, len(items))
    c.setFont("Helvetica", 10)
    c.drawString(44, table_bottom - 14, f"HST#: {inv['hst_number']}")
    c.setFont("Helvetica-Bold", 10)
    c.drawString(460, table_bottom - 10, "Total")
    c.setFont("Helvetica", 10)
    c.drawRightString(545, table_bottom - 10, fmt_money(inv["total"]))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(450, table_bottom - 28, "BALANCE")
    c.drawString(460, table_bottom - 41, "DUE")
    c.setFont("Helvetica", 10)
    c.drawRightString(545, table_bottom - 28, fmt_money(inv["total"]))
    c.drawRightString(545, table_bottom - 41, inv["currency"])
    c.setFont("Helvetica", 10)
    words = money_words(inv["total"])
    c.drawCentredString(width / 2, table_bottom - 68, words)
    c.save()


def make_template():
    with db() as conn:
        locations, customers, tasks, _, _, _, _ = get_master_maps(conn)
    wb = Workbook()
    ws = wb.active
    ws.title = "InvoiceUpload"
    ws.append(EXPECTED_HEADERS)
    for col in range(1, len(EXPECTED_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(14, len(cell.value) + 2)
    ws.append([
        dt.date.today().strftime("%d/%m/%Y"),
        dt.date.today().strftime("%Y%m"),
        locations[0]["code"] if locations else "",
        f"{customers[0]['external_code']} - {customers[0]['name']}" if customers else "",
        "Resource Name",
        tasks[0]["name"] if tasks else "Consulting",
        "CAD",
        100,
        160,
        "Hours",
        "",
        "-",
    ])

    master = wb.create_sheet("MasterData")
    master.append(["Locations", "Customers", "Tasks", "Currency", "InvoiceType"])
    for i, loc in enumerate(locations, start=2):
        master.cell(i, 1, loc["code"])
    for i, cust in enumerate(customers, start=2):
        master.cell(i, 2, f"{cust['external_code']} - {cust['name']}")
    for i, task in enumerate(tasks, start=2):
        master.cell(i, 3, task["name"])
    for i, value in enumerate(["CAD", "USD"], start=2):
        master.cell(i, 4, value)
    for i, value in enumerate(["Hours", "Fixed"], start=2):
        master.cell(i, 5, value)
    master.sheet_state = "hidden"

    validations = {
        "C": f"'MasterData'!$A$2:$A${max(2, len(locations) + 1)}",
        "D": f"'MasterData'!$B$2:$B${max(2, len(customers) + 1)}",
        "F": f"'MasterData'!$C$2:$C${max(2, len(tasks) + 1)}",
        "G": "'MasterData'!$D$2:$D$3",
        "J": "'MasterData'!$E$2:$E$3",
    }
    for col, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1000")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def html_page(title, body, user=None, active=""):
    nav = ""
    if user:
        links = [
            ("/", "Upload"),
            ("/history", "History"),
            ("/masters?type=customers", "Masters"),
            ("/users", "Users"),
            ("/logout", "Logout"),
        ]
        nav = "<nav>" + "".join(
            f'<a class="{ "active" if active == label else "" }" href="{href}">{label}</a>' for href, label in links
        ) + f"<span>{esc(user['username'])}</span></nav>"
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)} - {APP_NAME}</title>
  <link rel="stylesheet" href="/static/app.css">
</head>
<body>
  <header><div class="brand"><span class="mark">B</span><strong>Banking Labs</strong><small>Invoice Generation</small></div>{nav}</header>
  <main>{body}</main>
</body>
</html>"""


CSS = """
:root { color-scheme: light; --blue:#2f67c7; --ink:#1e2329; --muted:#667085; --line:#d9dee8; --bg:#f6f8fb; --panel:#fff; --ok:#0f766e; --bad:#b42318; }
* { box-sizing: border-box; }
body { margin:0; font-family: Arial, Helvetica, sans-serif; color:var(--ink); background:var(--bg); }
header { min-height:64px; display:flex; align-items:center; justify-content:space-between; padding:0 28px; background:#fff; border-bottom:1px solid var(--line); position:sticky; top:0; z-index:2; }
.brand { display:flex; align-items:center; gap:12px; }
.brand .mark { width:36px; height:36px; border-radius:6px; background:var(--blue); color:white; display:grid; place-items:center; font-size:27px; font-weight:700; }
.brand small { color:var(--muted); border-left:1px solid var(--line); padding-left:12px; }
nav { display:flex; align-items:center; gap:6px; }
nav a { color:var(--muted); text-decoration:none; padding:10px 12px; border-radius:6px; font-size:14px; }
nav a.active, nav a:hover { color:var(--blue); background:#eef4ff; }
nav span { color:var(--muted); margin-left:8px; font-size:13px; }
main { max-width:1180px; margin:0 auto; padding:28px; }
.grid { display:grid; grid-template-columns: minmax(320px, 1fr) minmax(360px, 1.4fr); gap:22px; align-items:start; }
.panel { background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:20px; box-shadow:0 1px 2px rgba(16,24,40,.04); }
h1 { font-size:24px; margin:0 0 18px; }
h2 { font-size:18px; margin:0 0 14px; }
h3 { font-size:15px; margin:18px 0 10px; }
p { color:var(--muted); line-height:1.45; }
label { display:block; font-size:13px; color:#344054; margin:12px 0 6px; }
input, select, textarea { width:100%; padding:10px 11px; border:1px solid #cfd6e3; border-radius:6px; font-size:14px; background:white; }
input[type=file] { padding:8px; }
button, .button { display:inline-flex; align-items:center; justify-content:center; min-height:38px; border:1px solid var(--blue); background:var(--blue); color:white; padding:8px 14px; border-radius:6px; font-weight:700; font-size:14px; text-decoration:none; cursor:pointer; }
button.secondary, .button.secondary { background:white; color:var(--blue); }
button.danger { background:#fff; color:var(--bad); border-color:#f1b4ad; }
.actions { display:flex; flex-wrap:wrap; gap:10px; align-items:center; margin-top:16px; }
.flash { border-radius:6px; padding:10px 12px; margin:0 0 16px; border:1px solid #b6ddff; background:#eff8ff; color:#175cd3; }
.flash.error { border-color:#fecdca; background:#fffbfa; color:var(--bad); }
.status { padding:4px 8px; border-radius:999px; font-size:12px; font-weight:700; background:#eef4ff; color:#175cd3; }
.status.failed { background:#fef3f2; color:var(--bad); }
.table-scroll { overflow-x:auto; }
table { width:100%; border-collapse:collapse; background:white; }
th, td { padding:10px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; font-size:14px; }
th { font-size:12px; text-transform:uppercase; color:#667085; background:#f9fafb; }
td.num, th.num { text-align:right; }
.muted { color:var(--muted); }
.tabs { display:flex; flex-wrap:wrap; gap:8px; margin-bottom:16px; }
.tabs a { text-decoration:none; color:#344054; border:1px solid var(--line); padding:8px 10px; border-radius:6px; background:#fff; }
.tabs a.active { color:var(--blue); border-color:#9ec5fe; background:#eef4ff; }
.row-form { display:grid; grid-template-columns: repeat(4, 1fr) auto; gap:8px; align-items:end; }
.wide { grid-column: span 2; }
.login { max-width:420px; margin:72px auto; }
.compact input { min-width:110px; }
@media (max-width: 850px) {
  header { align-items:flex-start; padding:14px; gap:12px; flex-direction:column; }
  nav { flex-wrap:wrap; }
  main { padding:18px; }
  .grid, .row-form { grid-template-columns:1fr; }
  .wide { grid-column:auto; }
}
"""


def route_dashboard(user, qs=None, flash=""):
    with db() as conn:
        batches = conn.execute(
            """
            SELECT b.*, u.username FROM upload_batches b JOIN users u ON u.id = b.uploaded_by
            ORDER BY b.uploaded_at DESC LIMIT 10
            """
        ).fetchall()
    rows = "".join(
        f"<tr><td>{esc(b['uploaded_at'])}</td><td>{esc(b['filename'])}</td><td><span class='status {esc(b['status'])}'>{esc(b['status'])}</span></td>"
        f"<td class='num'>{b['row_count']}</td><td class='num'>{b['invoice_count']}</td><td><a href='/batch?id={b['id']}'>Open</a></td></tr>"
        for b in batches
    )
    body = f"""
    {flash}
    <div class="grid">
      <section class="panel">
        <h1>Generate Invoices</h1>
        <form method="post" action="/upload" enctype="multipart/form-data">
          <label>Invoice Excel</label>
          <input type="file" name="excel" accept=".xlsx" required>
          <div class="actions">
            <button type="submit">Generate PDF</button>
            <a class="button secondary" href="/template.xlsx">Download Template</a>
          </div>
        </form>
        <h3>Required columns</h3>
        <p>{esc(", ".join(EXPECTED_HEADERS))}</p>
      </section>
      <section class="panel">
        <h2>Recent Uploads</h2>
        <div class="table-scroll"><table><thead><tr><th>Uploaded</th><th>File</th><th>Status</th><th class="num">Rows</th><th class="num">Invoices</th><th></th></tr></thead><tbody>{rows or '<tr><td colspan="6" class="muted">No uploads yet.</td></tr>'}</tbody></table></div>
      </section>
    </div>
    """
    return html_page("Upload", body, user, "Upload")


def route_history(user):
    with db() as conn:
        invoices = conn.execute(
            """
            SELECT i.*, b.filename, c.name AS customer_name, l.code AS location_code,
                   (SELECT GROUP_CONCAT(DISTINCT it.resource_name) FROM invoice_items it WHERE it.invoice_id = i.id) AS resource_names
            FROM invoices i
            JOIN upload_batches b ON b.id = i.batch_id
            JOIN customers c ON c.id = i.customer_id
            JOIN locations l ON l.id = i.location_id
            ORDER BY i.created_at DESC LIMIT 200
            """
        ).fetchall()
    rows = "".join(
        f"<tr><td>{inv['invoice_number']}</td><td>{esc(inv['created_at'])}</td><td>{esc(inv['customer_name'])}</td><td>{esc(inv['location_code'])}</td>"
        f"<td>{esc(inv['po_number'])}</td><td>{esc(inv['lto'])}</td><td>{esc(inv['resource_names'] or '')}</td><td class='num'>{fmt_money(inv['total'])} {esc(inv['currency'])}</td>"
        f"<td><a href='/invoice.pdf?id={inv['id']}'>PDF</a> · <a href='/batch?id={inv['batch_id']}'>Details</a></td></tr>"
        for inv in invoices
    )
    body = f"""
    <section class="panel">
      <h1>Invoice History</h1>
      <div class="table-scroll"><table><thead><tr><th>Invoice #</th><th>Created</th><th>Customer</th><th>Location</th><th>PO #</th><th>LTO</th><th>Resource</th><th class="num">Total</th><th></th></tr></thead><tbody>{rows or '<tr><td colspan="9" class="muted">No invoices generated yet.</td></tr>'}</tbody></table></div>
    </section>
    """
    return html_page("History", body, user, "History")


def route_batch(user, batch_id):
    with db() as conn:
        batch = conn.execute(
            "SELECT b.*, u.username FROM upload_batches b JOIN users u ON u.id = b.uploaded_by WHERE b.id = ?",
            (batch_id,),
        ).fetchone()
        if not batch:
            return None
        invoices = conn.execute(
            """
            SELECT i.*, c.name AS customer_name, l.code AS location_code,
                   (SELECT GROUP_CONCAT(DISTINCT it.resource_name) FROM invoice_items it WHERE it.invoice_id = i.id) AS resource_names
            FROM invoices i JOIN customers c ON c.id = i.customer_id JOIN locations l ON l.id = i.location_id
            WHERE i.batch_id = ? ORDER BY i.invoice_number
            """,
            (batch_id,),
        ).fetchall()
        items = conn.execute(
            """
            SELECT it.*, i.invoice_number
            FROM invoice_items it JOIN invoices i ON i.id = it.invoice_id
            WHERE i.batch_id = ? ORDER BY i.invoice_number, it.source_row
            """,
            (batch_id,),
        ).fetchall()
    errors = json.loads(batch["error_json"] or "[]")
    error_html = ""
    if errors:
        error_html = "<div class='flash error'><strong>Upload errors</strong><ul>" + "".join(
            f"<li>Row {e['row']}: {esc(e['error'])}</li>" for e in errors
        ) + "</ul></div>"
    inv_rows = "".join(
        f"<tr><td>{i['invoice_number']}</td><td>{esc(i['customer_name'])}</td><td>{esc(i['location_code'])}</td><td>{esc(i['po_number'])}</td><td>{esc(i['lto'])}</td><td>{esc(i['resource_names'] or '')}</td><td class='num'>{fmt_money(i['total'])}</td><td><a href='/invoice.pdf?id={i['id']}'>PDF</a></td></tr>"
        for i in invoices
    )
    item_rows = "".join(
        f"<tr><td>{it['source_row']}</td><td>{it['invoice_number']}</td><td>{esc(it['invoice_month'])}</td><td>{esc(it['activity'])}</td><td>{esc(it['resource_name'])}</td><td class='num'>{fmt_money(it['hours'])}</td><td class='num'>{fmt_money(it['rate'])}</td><td class='num'>{fmt_money(it['amount'])}</td></tr>"
        for it in items
    )
    body = f"""
    <section class="panel">
      <h1>Upload Batch</h1>
      {error_html}
      <p><strong>{esc(batch['filename'])}</strong> uploaded by {esc(batch['username'])} at {esc(batch['uploaded_at'])}. Status: <span class="status {esc(batch['status'])}">{esc(batch['status'])}</span></p>
      <div class="actions">
        <a class="button secondary" href="/batch-source?id={batch['id']}">Download Uploaded Excel</a>
        <a class="button secondary" href="/batch-zip?id={batch['id']}">Download PDFs ZIP</a>
      </div>
      <h2>Generated Invoices</h2>
      <div class="table-scroll"><table><thead><tr><th>Invoice #</th><th>Customer</th><th>Location</th><th>PO #</th><th>LTO</th><th>Resource</th><th class="num">Total</th><th></th></tr></thead><tbody>{inv_rows or '<tr><td colspan="8" class="muted">No invoices generated.</td></tr>'}</tbody></table></div>
      <h2>Uploaded Row Details</h2>
      <div class="table-scroll"><table><thead><tr><th>Source Row</th><th>Invoice #</th><th>Month</th><th>Activity</th><th>Resource</th><th class="num">Hours</th><th class="num">Rate</th><th class="num">Amount</th></tr></thead><tbody>{item_rows or '<tr><td colspan="8" class="muted">No parsed row details.</td></tr>'}</tbody></table></div>
    </section>
    """
    return html_page("Batch", body, user, "History")


MASTER_SCHEMAS = {
    "locations": ["code", "name", "address_line1", "address_line2", "phone", "email", "website", "hst_number"],
    "customers": ["external_code", "name", "address_line1", "address_line2", "city", "province_postal", "country", "attention", "default_po"],
    "tasks": ["name"],
    "tax_rates": ["code", "percent"],
}


def route_masters(user, table="customers", flash=""):
    table = table if table in MASTER_SCHEMAS else "customers"
    fields = MASTER_SCHEMAS[table]
    with db() as conn:
        rows = conn.execute(f"SELECT * FROM {table} ORDER BY id").fetchall()
    tabs = "".join(f'<a class="{ "active" if table == t else "" }" href="/masters?type={t}">{t.replace("_", " ").title()}</a>' for t in MASTER_SCHEMAS)
    header = "".join(f"<th>{esc(f.replace('_', ' ').title())}</th>" for f in fields)
    row_html = ""
    for row in rows:
        inputs = "".join(
            f"<td><input name='{esc(field)}' value='{esc(row[field])}'></td>" for field in fields
        )
        row_html += (
            f"<tr><form method='post' action='/master-save'><input type='hidden' name='type' value='{table}'><input type='hidden' name='id' value='{row['id']}'>"
            f"{inputs}<td><select name='active'><option value='1' {'selected' if row['active'] else ''}>Active</option><option value='0' {'selected' if not row['active'] else ''}>Inactive</option></select></td><td><button class='secondary'>Save</button></td></form></tr>"
        )
    add_inputs = "".join(f"<label>{esc(f.replace('_',' ').title())}<input name='{esc(f)}'></label>" for f in fields)
    body = f"""
    {flash}
    <section class="panel">
      <h1>Master Data</h1>
      <div class="tabs">{tabs}</div>
      <h2>{esc(table.replace('_', ' ').title())}</h2>
      <div class="table-scroll"><table class="compact"><thead><tr>{header}<th>Status</th><th></th></tr></thead><tbody>{row_html}</tbody></table></div>
      <h2>Add Record</h2>
      <form method="post" action="/master-add" class="row-form">
        <input type="hidden" name="type" value="{table}">
        {add_inputs}
        <button type="submit">Add</button>
      </form>
    </section>
    """
    return html_page("Masters", body, user, "Masters")


def route_users(user, flash=""):
    with db() as conn:
        users = conn.execute("SELECT id, username, role, active, created_at FROM users ORDER BY username").fetchall()
    rows = "".join(
        f"<tr><td>{esc(u['username'])}</td><td>{esc(u['role'])}</td><td>{'Active' if u['active'] else 'Inactive'}</td><td>{esc(u['created_at'])}</td></tr>"
        for u in users
    )
    body = f"""
    {flash}
    <section class="panel">
      <h1>Users</h1>
      <div class="table-scroll"><table><thead><tr><th>Username</th><th>Role</th><th>Status</th><th>Created</th></tr></thead><tbody>{rows}</tbody></table></div>
      <h2>Add User</h2>
      <form method="post" action="/user-add" class="row-form">
        <label>Username<input name="username" required></label>
        <label>Password<input name="password" type="password" minlength="8" required></label>
        <label>Role<select name="role"><option>admin</option></select></label>
        <button type="submit">Add</button>
      </form>
    </section>
    """
    return html_page("Users", body, user, "Users")


class App(BaseHTTPRequestHandler):
    server_version = "BLInvoice/1.0"

    def current_user(self):
        raw = self.headers.get("Cookie", "")
        jar = cookies.SimpleCookie(raw)
        token = jar.get("blinvoice_session")
        return verify_session(token.value if token else "")

    def send_bytes(self, data, content_type, filename=None, status=200):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(data)

    def send_html(self, html_text, status=200):
        self.send_bytes(html_text.encode("utf-8"), "text/html; charset=utf-8", status=status)

    def redirect(self, location, cookie_value=None, clear_cookie=False):
        self.send_response(303)
        self.send_header("Location", location)
        if cookie_value:
            self.send_header("Set-Cookie", f"blinvoice_session={cookie_value}; HttpOnly; SameSite=Lax; Path=/; Max-Age={SESSION_DAYS * 86400}")
        if clear_cookie:
            self.send_header("Set-Cookie", "blinvoice_session=; HttpOnly; SameSite=Lax; Path=/; Max-Age=0")
        self.end_headers()

    def parse_post(self):
        ctype, pdict = cgi.parse_header(self.headers.get("Content-Type", ""))
        if ctype == "multipart/form-data":
            pdict["boundary"] = pdict["boundary"].encode("utf-8")
            pdict["CONTENT-LENGTH"] = int(self.headers.get("Content-Length", 0))
            return cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"}, keep_blank_values=True)
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length).decode("utf-8")
        return {k: v[0] for k, v in parse_qs(raw, keep_blank_values=True).items()}

    def require_user(self):
        user = self.current_user()
        if not user:
            self.redirect("/login")
            return None
        return user

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)
        try:
            if path == "/static/app.css":
                return self.send_bytes(CSS.encode("utf-8"), "text/css; charset=utf-8")
            if path == "/login":
                body = """
                <section class="panel login">
                  <h1>Sign In</h1>
                  <form method="post" action="/login">
                    <label>Username</label><input name="username" required autofocus>
                    <label>Password</label><input type="password" name="password" required>
                    <div class="actions"><button type="submit">Sign In</button></div>
                  </form>
                  <p>Default first login: admin / admin123. Change it after deployment by adding a new admin user and disabling the default account in the database.</p>
                </section>
                """
                return self.send_html(html_page("Login", body))
            user = self.require_user()
            if not user:
                return
            if path == "/":
                return self.send_html(route_dashboard(user))
            if path == "/logout":
                return self.redirect("/login", clear_cookie=True)
            if path == "/template.xlsx":
                return self.send_bytes(make_template(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "BLInvoice_Template.xlsx")
            if path == "/history":
                return self.send_html(route_history(user))
            if path == "/batch":
                content = route_batch(user, int(qs.get("id", ["0"])[0]))
                return self.send_html(content if content else "Not found", status=200 if content else 404)
            if path == "/batch-source":
                return self.download_batch_source(int(qs.get("id", ["0"])[0]))
            if path == "/batch-zip":
                return self.download_batch_zip(int(qs.get("id", ["0"])[0]))
            if path == "/invoice.pdf":
                return self.download_invoice(int(qs.get("id", ["0"])[0]))
            if path == "/masters":
                return self.send_html(route_masters(user, qs.get("type", ["customers"])[0]))
            if path == "/users":
                return self.send_html(route_users(user))
            return self.send_html("Not found", status=404)
        except Exception:
            traceback.print_exc()
            return self.send_html("<h1>Server error</h1><pre>%s</pre>" % esc(traceback.format_exc()), status=500)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/login":
                form = self.parse_post()
                username = str(form.get("username", "")).strip()
                password = str(form.get("password", ""))
                with db() as conn:
                    user = conn.execute("SELECT * FROM users WHERE username = ? AND active = 1", (username,)).fetchone()
                if user and check_password(password, user["password_hash"]):
                    return self.redirect("/", sign_session(user["id"]))
                body = "<div class='flash error'>Invalid username or password.</div>"
                return self.send_html(html_page("Login", body + "<section class='panel login'><h1>Sign In</h1><form method='post' action='/login'><label>Username</label><input name='username' required autofocus><label>Password</label><input type='password' name='password' required><div class='actions'><button type='submit'>Sign In</button></div></form></section>"), status=401)
            user = self.require_user()
            if not user:
                return
            if path == "/upload":
                return self.handle_upload(user)
            form = self.parse_post()
            if path == "/master-add":
                return self.handle_master_add(user, form)
            if path == "/master-save":
                return self.handle_master_save(user, form)
            if path == "/user-add":
                return self.handle_user_add(user, form)
            return self.send_html("Not found", status=404)
        except Exception:
            traceback.print_exc()
            return self.send_html("<h1>Server error</h1><pre>%s</pre>" % esc(traceback.format_exc()), status=500)

    def handle_upload(self, user):
        form = self.parse_post()
        field = form["excel"] if "excel" in form else None
        if field is None or not field.filename:
            return self.send_html(route_dashboard(user, flash="<div class='flash error'>Choose an .xlsx file to upload.</div>"), status=400)
        temp = UPLOAD_DIR / f"tmp_{uuid.uuid4()}.xlsx"
        with temp.open("wb") as out:
            shutil.copyfileobj(field.file, out)
        try:
            batch_id, errors, invoice_ids = create_invoices_from_upload(temp, Path(field.filename).name, user["id"])
            if errors:
                return self.redirect(f"/batch?id={batch_id}")
            return self.redirect(f"/batch?id={batch_id}")
        finally:
            if temp.exists():
                temp.unlink()

    def handle_master_add(self, user, form):
        table = form.get("type", "")
        if table not in MASTER_SCHEMAS:
            return self.send_html("Bad master type", status=400)
        fields = MASTER_SCHEMAS[table]
        values = [str(form.get(f, "")).strip() for f in fields]
        with db() as conn:
            placeholders = ", ".join("?" for _ in fields)
            conn.execute(f"INSERT INTO {table} ({', '.join(fields)}, active) VALUES ({placeholders}, 1)", values)
        return self.redirect(f"/masters?type={table}")

    def handle_master_save(self, user, form):
        table = form.get("type", "")
        if table not in MASTER_SCHEMAS:
            return self.send_html("Bad master type", status=400)
        row_id = int(form.get("id", "0"))
        fields = MASTER_SCHEMAS[table]
        values = [str(form.get(f, "")).strip() for f in fields]
        active = 1 if form.get("active") == "1" else 0
        sets = ", ".join(f"{f} = ?" for f in fields) + ", active = ?"
        with db() as conn:
            conn.execute(f"UPDATE {table} SET {sets} WHERE id = ?", values + [active, row_id])
        return self.redirect(f"/masters?type={table}")

    def handle_user_add(self, user, form):
        username = str(form.get("username", "")).strip()
        password = str(form.get("password", ""))
        role = str(form.get("role", "admin")).strip() or "admin"
        if len(password) < 8:
            return self.send_html(route_users(user, "<div class='flash error'>Password must be at least 8 characters.</div>"), status=400)
        with db() as conn:
            conn.execute(
                "INSERT INTO users (username, password_hash, role, active, created_at) VALUES (?, ?, ?, 1, ?)",
                (username, pbkdf2_hash(password), role, utcnow()),
            )
        return self.redirect("/users")

    def download_invoice(self, invoice_id):
        with db() as conn:
            inv = conn.execute(
                """
                SELECT i.*,
                       (SELECT GROUP_CONCAT(DISTINCT it.resource_name) FROM invoice_items it WHERE it.invoice_id = i.id) AS resource_names,
                       (SELECT it.invoice_month FROM invoice_items it WHERE it.invoice_id = i.id LIMIT 1) AS invoice_month
                FROM invoices i WHERE i.id = ?
                """,
                (invoice_id,),
            ).fetchone()
        if not inv:
            return self.send_html("Not found", status=404)
        path = Path(inv["pdf_path"])
        filename = invoice_pdf_filename(inv["resource_names"], inv["invoice_month"], inv["invoice_number"])
        return self.send_bytes(path.read_bytes(), "application/pdf", filename)

    def download_batch_source(self, batch_id):
        with db() as conn:
            batch = conn.execute("SELECT * FROM upload_batches WHERE id = ?", (batch_id,)).fetchone()
        if not batch:
            return self.send_html("Not found", status=404)
        return self.send_bytes(Path(batch["stored_path"]).read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", batch["filename"])

    def download_batch_zip(self, batch_id):
        with db() as conn:
            rows = conn.execute(
                """
                SELECT i.invoice_number, i.pdf_path,
                       (SELECT GROUP_CONCAT(DISTINCT it.resource_name) FROM invoice_items it WHERE it.invoice_id = i.id) AS resource_names,
                       (SELECT it.invoice_month FROM invoice_items it WHERE it.invoice_id = i.id LIMIT 1) AS invoice_month
                FROM invoices i WHERE i.batch_id = ? ORDER BY i.invoice_number
                """,
                (batch_id,),
            ).fetchall()
        out = io.BytesIO()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
            for row in rows:
                path = Path(row["pdf_path"])
                if path.exists():
                    filename = invoice_pdf_filename(row["resource_names"], row["invoice_month"], row["invoice_number"])
                    zf.write(path, filename)
        return self.send_bytes(out.getvalue(), "application/zip", f"Batch_{batch_id}_Invoices.zip")


def main():
    init_db()
    host = os.environ.get("BLINVOICE_HOST", "127.0.0.1")
    port = int(os.environ.get("BLINVOICE_PORT", "8000"))
    httpd = ThreadingHTTPServer((host, port), App)
    print(f"{APP_NAME} running at http://{host}:{port}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass


if __name__ == "__main__":
    main()
